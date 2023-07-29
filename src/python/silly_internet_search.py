# #!/usr/bin/env python3

# meh
import openpyxl
import json
from openpyxl import load_workbook


def get_sheet_dict(ws):
    year_dict = {}
    column_headers = {}

    first_row = ws[1]

    for cell in first_row[1:]:
        column_headers[cell.column] = cell.value

    for row in ws.iter_rows(min_row=2):
        month = row[0].value
        temp_cell_dict = {}
        for cell in row[1:]:
            col_name = column_headers[cell.column]
            temp_cell_dict[col_name] = cell.value

        year_dict[month] = temp_cell_dict

    return year_dict


#############################################
#  MAIN
#############################################
if __name__ == '__main__':

    wb = load_workbook(
        filename="E:\GDrive\Harpers Startup\Excellent-Development\src\python\single_test.xlsx",
        read_only=False,
        data_only=True)  # Changes here

    budget_dict = {}

    for ws in wb:
        for row in ws.iter_rows(max_row=5):  # print first 5 rows
            print([cell.value for cell in row])
            
        year_dict = get_sheet_dict(ws)
        budget_dict[ws.title] = year_dict

    data_json = json.dumps(budget_dict, indent=4)  # pretty print JSON

    with open("new.json", "w") as f:
        f.write(data_json)
    print(data_json)



# GETTING THERE
# import pandas as pd
# import json
# from collections import defaultdict
# import fuckit

# path = "E:\GDrive\Harpers Startup\Excellent-Development\src\python\single_test.xlsx"
# df = pd.read_excel(path,
#                    engine='openpyxl',
#                    header=None, names=['Block', 'Task', 'Code']  # only if your file has no headers
#                    )

# df.dropna(inplace=True, axis=0, how='all')
# df.fillna(method='ffill', inplace=True, axis=0)
# df = df.set_index(['Block', 'Task'])

# nested_dict = defaultdict(lambda : defaultdict(list))

# for keys, value in df.Code.iteritems():
#     nested_dict[keys[0]][keys[1]].append(value)

# json_str = json.dumps(nested_dict, indent=4)
# print(json_str)


# ABSOLUTELY FUCKED
# import pandas

# excel_data_df = pandas.read_excel('E:\GDrive\Harpers Startup\Excellent-Development\src\python\single_test.xlsx')

# json_str = excel_data_df.to_json()

# print('Excel Sheet to JSON:\n', json_str)

# with open("new.json", "w") as f:
#     f.write(json_str)
