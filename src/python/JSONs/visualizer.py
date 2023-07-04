import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
import psutil


def close_excel():
    for proc in psutil.process_iter():
        try:
            if proc.name() == "excel.exe":
                proc.kill()
                print("Excel process terminated successfully.")
                return
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass

    print("Excel process not found.")


def label_visualize():
    # Read data from JSON file
    with open('annotated_output.json', 'r') as json_file:
        data = json.load(json_file)

    # Create a new workbook and select the active sheet
    workbook = Workbook()

    # Define color codes for different annotations
    color_mapping = {
        'LABEL': '00FF00',    # Green
        'DATA': 'FF0000',     # Red
        'FORMULA': 'FFA500',  # Orange
    }

    # Process data and apply formatting
    for item in data:
        sheet_number = item['Sheet number']
        annotations = item['Annotations']

        # Create a new sheet based on the sheet number
        sheet = workbook.create_sheet(title=f"Sheet {sheet_number}")

        for cell_location, cell_info in annotations.items():
            value = cell_info['value']
            annotation = cell_info['annotation']

            cell = sheet.cell(
                row=int(cell_location[1:]), column=ord(cell_location[0]) - 64)
            cell.value = value

            if annotation in color_mapping:
                color = color_mapping[annotation]
                fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type='solid')
                cell.fill = fill

    # Remove the default sheet created by Workbook
    workbook.remove(workbook['Sheet'])

    # Save the workbook as an Excel file
    workbook.save('visuals/label_visualized.xlsx')


def solid_block_visualize():
    # Define color patterns
    green_fill = PatternFill(
        start_color="00FF00",
        end_color="00FF00",
        fill_type="solid")
    red_fill = PatternFill(
        start_color="FF0000",
        end_color="FF0000",
        fill_type="solid")
    orange_fill = PatternFill(
        start_color="FFA500",
        end_color="FFA500",
        fill_type="solid")

    # Load JSON data
    with open('blocks_output.json', 'r') as f:
        data = json.load(f)

    # Function to add data to a worksheet
    def add_data_to_sheet(ws, block):
        for cell, details in block['cells'].items():
            row, col = int(cell[1:]), ord(cell[0]) - ord('A') + 1
            cell = ws.cell(row=row, column=col, value=details['value'])
            cell.fill = green_fill if details['annotation'] == 'LABEL' else (
                red_fill if details['annotation'] == 'DATA' else orange_fill)

    # Create a new Workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet created

    # Create a new sheet for each sheet in the JSON data
    for sheet in data:
        ws_sheet = wb.create_sheet(title=f"Sheet {sheet['sheet_number']}")
        for annotation_type, blocks in sheet['blocks'].items():
            for i, block in enumerate(blocks, start=1):
                ws = wb.create_sheet(
                    title=f"Sheet {sheet['sheet_number']} {annotation_type} {i}")
                add_data_to_sheet(ws, block)
                # Copy all data to the sheet created for each sheet
                for row in ws.iter_rows():
                    for cell in row:
                        ws_sheet[cell.coordinate].value = cell.value
                        if cell.fill == green_fill:
                            ws_sheet[cell.coordinate].fill = green_fill
                        elif cell.fill == red_fill:
                            ws_sheet[cell.coordinate].fill = red_fill
                        elif cell.fill == orange_fill:
                            ws_sheet[cell.coordinate].fill = orange_fill

    # Save the workbook
    filename = 'visuals/solid_block_visualized.xlsx'
    wb.save(filename=filename)

    # Open the Excel file
    os.system('start excel.exe "{}"'.format(
        'visuals/solid_block_visualized.xlsx'))


def solid_table_visualize():
    # Define color patterns
    green_fill = PatternFill(
        start_color="00FF00",
        end_color="00FF00",
        fill_type="solid")
    red_fill = PatternFill(
        start_color="FF0000",
        end_color="FF0000",
        fill_type="solid")

    # Load JSON data
    with open('solid_table_output.json', 'r') as f:
        data = json.load(f)

    # Function to add data to a worksheet
    def add_data_to_sheet(ws, data):
        # Add data
        for key, value in data.items():
            cell, annotation = key, value['value']
            row, col = int(cell[1:]), ord(cell[0]) - ord('A') + 1
            cell = ws.cell(row=row, column=col, value=annotation)
            cell.fill = green_fill if value['annotation'] == 'LABEL' else red_fill

    # Create a new file for each table
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet created

    for sheet, value in data.items():
        full_ws = wb.create_sheet(title=f"Sheet {sheet} Full")

        free_solid_blocks = value['free_solid_blocks']
        if free_solid_blocks:
            if free_solid_blocks.get('LABEL'):
                for label in free_solid_blocks['LABEL']:
                    add_data_to_sheet(full_ws, label['cells'])
            if free_solid_blocks.get('DATA'):
                for data_block in free_solid_blocks['DATA']:
                    add_data_to_sheet(full_ws, data_block['cells'])

        if value.get('solid_tables'):
            for i, table in enumerate(value['solid_tables']):
                add_data_to_sheet(full_ws, table['data']['cells'])
                if table.get('labels'):
                    for place, label_data in table['labels'].items():
                        add_data_to_sheet(full_ws, label_data['cells'])

        if free_solid_blocks:
            ws = wb.create_sheet(title=f"Sheet {sheet} Free Solid Blocks")
            if free_solid_blocks.get('LABEL'):
                for label in free_solid_blocks['LABEL']:
                    add_data_to_sheet(ws, label['cells'])
            if free_solid_blocks.get('DATA'):
                for data_block in free_solid_blocks['DATA']:
                    add_data_to_sheet(ws, data_block['cells'])

        if value.get('solid_tables'):
            for i, table in enumerate(value['solid_tables']):
                ws = wb.create_sheet(title=f"Sheet {sheet} Table {i}")
                add_data_to_sheet(ws, table['data']['cells'])
                if table.get('labels'):
                    for place, label_data in table['labels'].items():
                        add_data_to_sheet(ws, label_data['cells'])

    wb.save(filename='visuals/output_table_visualized_sheet.xlsx')

    # Open the spreadsheet in Excel
    os.system('start excel.exe "{}"'.format(
        'visuals/output_table_visualized_sheet.xlsx'))


# Call the function to close previou Excel
close_excel()
label_visualize()
solid_block_visualize()
solid_table_visualize()
